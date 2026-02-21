"""
muhurtha_excel.py
Generates a formatted .xlsx workbook from monthly Muhurtha scan results.
"""

import io
import calendar
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

# ── Colour palette ────────────────────────────────────────────────────────────
GOLD_DARK   = "B8860B"   # dark gold — headers
GOLD_MID    = "C9A84C"   # mid gold — sub-headers
GOLD_LIGHT  = "F5DFA0"   # light gold — title text
INK         = "0E0C0A"   # near-black background
INK_LIGHT   = "1C1A16"   # slightly lighter — row backgrounds
PARCHMENT   = "F5EDD8"   # cream — body text
SCORE5_BG   = "2A2400"   # dark gold tint — excellent row bg
SCORE5_FG   = "FFD966"   # yellow-gold — excellent text
SCORE4_BG   = "1A2200"   # dark green tint — very good row bg
SCORE4_FG   = "A9C46C"   # green — very good text
BORDER_CLR  = "3A3020"   # subtle border
WHITE       = "FFFFFF"
GRAY_LIGHT  = "E8E0CC"

# ── Style helpers ──────────────────────────────────────────────────────────────

def _font(bold=False, size=10, color=WHITE, italic=False, name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(color=BORDER_CLR):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set_cell(cell, value, font=None, fill=None, align=None, border=None, fmt=None):
    cell.value = value
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    if fmt:    cell.number_format = fmt


# ── Cover Sheet ───────────────────────────────────────────────────────────────

def build_cover(wb, month_name, year, place, tz, custom_planets, results):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 3

    # Title block rows 1-8
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 10
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 10

    # Title background band
    for row in range(2, 5):
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = _fill(INK)

    # Main title
    ws.merge_cells("B2:E2")
    c = ws["B2"]
    _set_cell(c, f"Muhurtha Calendar — {month_name} {year}",
              font=_font(bold=True, size=20, color=GOLD_LIGHT, name="Arial"),
              fill=_fill(INK),
              align=_align("center", "center"))

    ws.merge_cells("B3:E3")
    c = ws["B3"]
    _set_cell(c, "Auspicious Timing Report · KP Astrology · Transit Chart",
              font=_font(size=10, color=GOLD_MID, italic=True),
              fill=_fill(INK),
              align=_align("center", "center"))

    # Meta info
    meta = [
        ("📍 Location",  place),
        ("🕐 Timezone",  tz),
        ("📅 Period",    f"{month_name} {year}"),
        ("⭐ Criteria",  "Excellent (🌟) & Very Good (✨) only"),
        ("⏱ Interval",  "2-minute scan steps"),
    ]
    if custom_planets:
        meta.append(("⚙️ Custom Planets", ", ".join(custom_planets)))

    row = 6
    for label, value in meta:
        ws.row_dimensions[row].height = 18
        c_lbl = ws.cell(row=row, column=2)
        c_val = ws.cell(row=row, column=3)
        _set_cell(c_lbl, label,
                  font=_font(bold=True, size=10, color=GOLD_MID),
                  align=_align("left", "center"))
        ws.merge_cells(f"C{row}:E{row}")
        _set_cell(c_val, value,
                  font=_font(size=10, color=PARCHMENT),
                  align=_align("left", "center"))
        row += 1

    # Legend
    row += 1
    ws.row_dimensions[row].height = 14
    ws.merge_cells(f"B{row}:E{row}")
    _set_cell(ws.cell(row=row, column=2), "LEGEND",
              font=_font(bold=True, size=9, color=GOLD_DARK),
              align=_align("left", "center"))
    row += 1

    legend_items = [
        (SCORE5_BG, SCORE5_FG, "🌟 Excellent Muhurtha",  "All 5 rules pass · Rule 5 = Kendra (arc 1/4/7/10)"),
        (SCORE4_BG, SCORE4_FG, "✨ Very Good Muhurtha",  "All 5 rules pass · Rule 5 = Just OK"),
    ]
    for bg, fg, grade, desc in legend_items:
        ws.row_dimensions[row].height = 18
        c1 = ws.cell(row=row, column=2)
        c2 = ws.cell(row=row, column=3)
        _set_cell(c1, grade,
                  font=_font(bold=True, size=10, color=fg),
                  fill=_fill(bg),
                  align=_align("left", "center"),
                  border=_border())
        ws.merge_cells(f"C{row}:E{row}")
        _set_cell(c2, desc,
                  font=_font(size=9, color=PARCHMENT),
                  fill=_fill(INK_LIGHT),
                  align=_align("left", "center"),
                  border=_border())
        row += 1

    # Summary table header
    row += 2
    ws.row_dimensions[row].height = 22
    headers = ["Activity", "Karaka Planets", "🌟 Excellent", "✨ Very Good", "Total Slots"]
    for i, h in enumerate(headers, 2):
        c = ws.cell(row=row, column=i)
        _set_cell(c, h,
                  font=_font(bold=True, size=10, color=GOLD_LIGHT),
                  fill=_fill(GOLD_DARK),
                  align=_align("center", "center"),
                  border=_border())
    row += 1

    # Summary rows — one per activity
    ACTIVITY_PLANETS = {
        "Travel":               "Moon, Mercury",
        "Finance / Investment": "Jupiter, Venus",
        "Job Application":      "Sun, Mercury",
        "Hard Work / Toil":     "Saturn",
        "Construction":         "Saturn, Mars",
        "Marriage":             "Venus",
        "Health / Recovery":    "Sun, Moon",
        "Legal Matters":        "Saturn, Mars",
        "Education / Study":    "Mercury, Jupiter",
        "Custom":               ", ".join(custom_planets) if custom_planets else "—",
    }

    ICONS = {
        "Travel": "✈️", "Finance / Investment": "💰", "Job Application": "💼",
        "Hard Work / Toil": "⚒️", "Construction": "🏗️", "Marriage": "💍",
        "Health / Recovery": "🏥", "Legal Matters": "⚖️", "Education / Study": "📚",
        "Custom": "⚙️",
    }

    alt = False
    for label, slots in results.items():
        if label == "Custom" and not custom_planets:
            continue
        ws.row_dimensions[row].height = 18
        bg = "161410" if alt else INK_LIGHT
        excellent = sum(1 for w in slots if w["best_score"] == 5)
        very_good = sum(1 for w in slots if w["best_score"] == 4)
        total = len(slots)
        planets = ACTIVITY_PLANETS.get(label, "")
        icon = ICONS.get(label, "")

        row_data = [f"{icon} {label}", planets, excellent, very_good, total]
        for i, val in enumerate(row_data, 2):
            c = ws.cell(row=row, column=i)
            is_number = isinstance(val, int)
            _set_cell(c, val,
                      font=_font(size=10, color=SCORE5_FG if (i == 4 and val > 0) else
                                              SCORE4_FG if (i == 5 and val > 0) else PARCHMENT),
                      fill=_fill(bg),
                      align=_align("center" if is_number else "left", "center"),
                      border=_border())
        alt = not alt
        row += 1

    # Sheet link hints
    row += 2
    ws.row_dimensions[row].height = 14
    ws.merge_cells(f"B{row}:E{row}")
    _set_cell(ws.cell(row=row, column=2),
              "Each activity has its own sheet with full slot details →",
              font=_font(size=9, color=GOLD_MID, italic=True),
              align=_align("left", "center"))


# ── Activity Sheet ────────────────────────────────────────────────────────────

def build_activity_sheet(wb, label, icon, planets, slots, month_name, year, tz):
    safe_name = label.replace("/", "-").replace(":", "")[:31]
    ws = wb.create_sheet(title=f"{icon} {safe_name}"[:31])
    ws.sheet_view.showGridLines = False

    # Column widths
    col_widths = [3, 14, 12, 12, 14, 22, 14, 18, 3]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet title ──
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 10

    for row in range(2, 5):
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = _fill(INK)

    ws.merge_cells("B2:H2")
    _set_cell(ws["B2"], f"{icon}  {label}  —  {month_name} {year}",
              font=_font(bold=True, size=16, color=GOLD_LIGHT),
              fill=_fill(INK),
              align=_align("left", "center"))

    ws.merge_cells("B3:H3")
    subtitle = f"Karaka Planets: {planets}   ·   Timezone: {tz}   ·   2-min scan intervals"
    _set_cell(ws["B3"], subtitle,
              font=_font(size=9, color=GOLD_MID, italic=True),
              fill=_fill(INK),
              align=_align("left", "center"))

    ws.merge_cells("B4:H4")
    criteria = "Showing: 🌟 Excellent Muhurtha (score 5)  &  ✨ Very Good Muhurtha (score 4)  —  Top 5 per month"
    _set_cell(ws["B4"], criteria,
              font=_font(size=9, color=PARCHMENT),
              fill=_fill(INK),
              align=_align("left", "center"))

    # ── Column headers ──
    row = 6
    ws.row_dimensions[row].height = 22
    headers = ["#", "Date", "Start Time", "End Time", "Duration", "Grade", "Lagna", "Rules Passed"]
    col_aligns = ["center","left","center","center","center","left","center","center"]
    for i, (h, ha) in enumerate(zip(headers, col_aligns), 2):
        c = ws.cell(row=row, column=i)
        _set_cell(c, h,
                  font=_font(bold=True, size=10, color=GOLD_LIGHT),
                  fill=_fill(GOLD_DARK),
                  align=_align(ha, "center"),
                  border=_border())
    row += 1

    # ── Data rows ──
    if not slots:
        ws.row_dimensions[row].height = 24
        ws.merge_cells(f"B{row}:H{row}")
        _set_cell(ws.cell(row=row, column=2),
                  "No Excellent or Very Good Muhurtha windows found this month.",
                  font=_font(size=10, color=PARCHMENT, italic=True),
                  fill=_fill(INK_LIGHT),
                  align=_align("center", "center"),
                  border=_border())
        return

    for rank, w in enumerate(slots, 1):
        ws.row_dimensions[row].height = 22
        score = w["best_score"]
        bg = SCORE5_BG if score == 5 else SCORE4_BG
        fg = SCORE5_FG if score == 5 else SCORE4_FG

        row_vals = [
            rank,
            w["date_str"],
            w["start_str"],
            w["end_str"],
            w["duration"],
            w["grade"],
            w["asc_sign"],
            f"{w['rules_pass']} / 5",
        ]
        col_aligns = ["center","left","center","center","center","left","center","center"]
        for i, (val, ha) in enumerate(zip(row_vals, col_aligns), 2):
            c = ws.cell(row=row, column=i)
            text_color = fg if i in (2, 7) else PARCHMENT
            if i == 7:  # Grade column — use accent color
                text_color = fg
            _set_cell(c, val,
                      font=_font(size=10, color=text_color,
                                 bold=(i == 2)),
                      fill=_fill(bg),
                      align=_align(ha, "center"),
                      border=_border())
        row += 1

    # ── Summary footer ──
    row += 1
    ws.row_dimensions[row].height = 18
    excellent_count = sum(1 for w in slots if w["best_score"] == 5)
    very_good_count = sum(1 for w in slots if w["best_score"] == 4)

    ws.merge_cells(f"B{row}:D{row}")
    _set_cell(ws.cell(row=row, column=2),
              f"🌟 Excellent: {excellent_count}   ·   ✨ Very Good: {very_good_count}   ·   Total: {len(slots)}",
              font=_font(bold=True, size=10, color=GOLD_MID),
              fill=_fill(INK),
              align=_align("left", "center"))

    # ── Rule reference box ──
    row += 2
    rule_info = [
        ("MUHURTHA RULE REFERENCE", None, True),
        ("Rule 1", "Lagna lord in kendra (1/4/7/10) or 11th house from transit lagna", False),
        ("Rule 2", "3rd house lord in kendra or 11th house from transit lagna", False),
        ("Rule 3", "11th house lord in kendra or 11th house from transit lagna", False),
        ("Rule 4", f"Karaka planet(s) [{planets}] in kendra or 11th from lagna", False),
        ("Rule 5", "Lagna lord & Lagna nakshatra lord: Kendra arc = Excellent, 2/3/11/12 = OK, 5/6/8/9 = Avoid", False),
        ("Override", "Ketu in Lagna → All grades nullified (Avoid)", False),
    ]
    for label_r, desc, is_header in rule_info:
        ws.row_dimensions[row].height = 16
        if is_header:
            ws.merge_cells(f"B{row}:H{row}")
            _set_cell(ws.cell(row=row, column=2), label_r,
                      font=_font(bold=True, size=9, color=GOLD_DARK),
                      fill=_fill(INK),
                      align=_align("left", "center"))
        else:
            c_lbl = ws.cell(row=row, column=2)
            _set_cell(c_lbl, label_r,
                      font=_font(bold=True, size=9, color=GOLD_MID),
                      fill=_fill(INK),
                      align=_align("left", "center"))
            ws.merge_cells(f"C{row}:H{row}")
            _set_cell(ws.cell(row=row, column=3), desc,
                      font=_font(size=9, color=PARCHMENT),
                      fill=_fill(INK),
                      align=_align("left", "center"))
        row += 1


# ── Master Sheet (all activities combined) ────────────────────────────────────

def build_master_sheet(wb, results, month_name, year, tz, custom_planets):
    ws = wb.create_sheet(title="📋 All Windows", index=1)
    ws.sheet_view.showGridLines = False

    col_widths = [3, 22, 14, 12, 12, 14, 22, 14, 3]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 14

    for col in range(1, 10):
        ws.cell(row=2, column=col).fill = _fill(INK)

    ws.merge_cells("B2:H2")
    _set_cell(ws["B2"], f"All Muhurtha Windows — {month_name} {year}",
              font=_font(bold=True, size=15, color=GOLD_LIGHT),
              fill=_fill(INK),
              align=_align("left", "center"))

    # Header
    row = 4
    ws.row_dimensions[row].height = 22
    headers = ["Activity", "Date", "Start", "End", "Duration", "Grade", "Lagna", "Score"]
    for i, h in enumerate(headers, 2):
        _set_cell(ws.cell(row=row, column=i), h,
                  font=_font(bold=True, size=10, color=GOLD_LIGHT),
                  fill=_fill(GOLD_DARK),
                  align=_align("center", "center"),
                  border=_border())
    row += 1

    ICONS = {
        "Travel": "✈️", "Finance / Investment": "💰", "Job Application": "💼",
        "Hard Work / Toil": "⚒️", "Construction": "🏗️", "Marriage": "💍",
        "Health / Recovery": "🏥", "Legal Matters": "⚖️", "Education / Study": "📚",
        "Custom": "⚙️",
    }

    # Collect all windows, sort by date then score
    all_windows = []
    for label, slots in results.items():
        if label == "Custom" and not custom_planets:
            continue
        icon = ICONS.get(label, "")
        for w in slots:
            all_windows.append((label, icon, w))

    all_windows.sort(key=lambda x: (x[2]["date_full"], -x[2]["best_score"]))

    alt = False
    for label, icon, w in all_windows:
        ws.row_dimensions[row].height = 20
        score = w["best_score"]
        bg = SCORE5_BG if score == 5 else (SCORE4_BG if score == 4 else ("161410" if alt else INK_LIGHT))
        fg = SCORE5_FG if score == 5 else (SCORE4_FG if score == 4 else PARCHMENT)

        row_vals = [
            f"{icon} {label}",
            w["date_str"],
            w["start_str"],
            w["end_str"],
            w["duration"],
            w["grade"],
            w["asc_sign"],
            "🌟" if score == 5 else "✨",
        ]
        col_aligns = ["left","left","center","center","center","left","center","center"]
        for i, (val, ha) in enumerate(zip(row_vals, col_aligns), 2):
            _set_cell(ws.cell(row=row, column=i), val,
                      font=_font(size=10, color=fg if i in (2, 8) else PARCHMENT,
                                 bold=(i == 2)),
                      fill=_fill(bg),
                      align=_align(ha, "center"),
                      border=_border())
        alt = not alt
        row += 1

    if not all_windows:
        ws.merge_cells(f"B{row}:H{row}")
        _set_cell(ws.cell(row=row, column=2),
                  "No Excellent or Very Good windows found this month.",
                  font=_font(size=10, color=PARCHMENT, italic=True),
                  fill=_fill(INK_LIGHT),
                  align=_align("center", "center"))


# ── Main export function ──────────────────────────────────────────────────────

def generate_excel(results, month_name, year, place, tz, custom_planets):
    """
    Returns a BytesIO stream containing the formatted .xlsx workbook.
    """
    ACTIVITY_PLANETS = {
        "Travel":               "Moon, Mercury",
        "Finance / Investment": "Jupiter, Venus",
        "Job Application":      "Sun, Mercury",
        "Hard Work / Toil":     "Saturn",
        "Construction":         "Saturn, Mars",
        "Marriage":             "Venus",
        "Health / Recovery":    "Sun, Moon",
        "Legal Matters":        "Saturn, Mars",
        "Education / Study":    "Mercury, Jupiter",
        "Custom":               ", ".join(custom_planets) if custom_planets else "",
    }
    ICONS = {
        "Travel": "✈️", "Finance / Investment": "💰", "Job Application": "💼",
        "Hard Work / Toil": "⚒️", "Construction": "🏗️", "Marriage": "💍",
        "Health / Recovery": "🏥", "Legal Matters": "⚖️", "Education / Study": "📚",
        "Custom": "⚙️",
    }

    wb = Workbook()

    # Cover sheet (uses wb.active)
    build_cover(wb, month_name, year, place, tz, custom_planets, results)

    # Master "all windows" sheet
    build_master_sheet(wb, results, month_name, year, tz, custom_planets)

    # One sheet per activity
    for label, slots in results.items():
        if label == "Custom" and not custom_planets:
            continue
        icon = ICONS.get(label, "⚙️")
        planets = ACTIVITY_PLANETS.get(label, "")
        build_activity_sheet(wb, label, icon, planets, slots, month_name, year, tz)

    # Save to BytesIO
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
